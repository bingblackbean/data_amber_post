import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
import schedule
import time
import locale
locale.setlocale(locale.LC_ALL, 'zh')


def report_data_gen(df, calc_date):
    # generate the data for report
    def price_statistic(df, calc_date):
        calc_df = df[df['Date'] == calc_date]
        mean_price = calc_df['AveragePrice'].mean()
        num_store = len(calc_df['region'].unique())
        sum_sale = calc_df['Total Volume'].sum()
        high_price = calc_df['AveragePrice'].max()
        high_store = calc_df.loc[calc_df['AveragePrice'].idxmax()]['region']
        high_store_sale = calc_df.loc[calc_df['AveragePrice'].idxmax(
        )]['Total Volume']
        low_price = calc_df['AveragePrice'].min()
        low_store = calc_df.loc[calc_df['AveragePrice'].idxmin()]['region']
        low_store_sale = calc_df.loc[calc_df['AveragePrice'].idxmin(
        )]['Total Volume']
        return [
            mean_price,
            num_store,
            sum_sale,
            high_price,
            high_store,
            high_store_sale,
            low_price,
            low_store,
            low_store_sale]

    calc_date = pd.to_datetime(str(calc_date))  # convert timestamp
    calc_date_str = calc_date.strftime('%Y年%m月%d日')  # printf as string
    ststs = price_statistic(df, calc_date)  # calculate statistics
    return [calc_date_str] + ststs


def report_sheet_update(report_sheet, external_data, raw_data_sheet):
    to_fill_cells_list = ['B2',
                          'C12', 'E12', 'H12',
                          'C14', 'E14', 'H14',
                          'C16', 'E16', 'H16'
                          ]  # define totill cell list
    for i, c in enumerate(to_fill_cells_list):
        report_sheet[c] = external_data[i]

    def create_bar_chart(chart_sheet, data_sheet):
        bar_chart = BarChart()
        bar_chart.title = 'TOP10 门店信息'
        row_start = 1
        row_end = 12 + row_start
        y = Reference(
            data_sheet,
            min_col=3,
            min_row=row_start,
            max_col=3,
            max_row=row_end)
        x = Reference(
            data_sheet,
            min_col=2,
            min_row=row_start,
            max_col=2,
            max_row=row_end)
        bar_chart.add_data(y, titles_from_data=True)
        bar_chart.set_categories(x)
        bar_chart.x_axis.title = '门店'
        bar_chart.y_axis.title = '价格'

        # Create a second chart
        line_chart = LineChart()
        y2 = Reference(
            data_sheet,
            min_col=4,
            min_row=row_start,
            max_col=4,
            max_row=row_end)

        line_chart.add_data(y2, titles_from_data=True)
        line_chart.y_axis.title = "销量"
        line_chart.y_axis.axId = 1
        gridline_sgp = GraphicalProperties(ln=LineProperties(noFill=True))
        line_chart.y_axis.majorGridlines.spPr = gridline_sgp
        bar_chart.y_axis.crosses = "max"

        # combine and adjust style
        bar_chart += line_chart
        bar_chart.legend.position = "t"
        bar_chart.height = 10  # default is 7.5
        bar_chart.width = 12  # default is 15
        props = GraphicalProperties(solidFill="f3f3f3")
        bar_chart.graphical_properties = props
        bar_chart.plot_area.graphicalProperties = props
        bar_chart.style = 2  # fast change color
        bar_chart.y_axis.majorGridlines.spPr = gridline_sgp
        chart_sheet.add_chart(bar_chart, 'B24')

    create_bar_chart(report_sheet, raw_data_sheet)


def raw_sheet_update(raw_data_sheet, df, calc_date):
    # update the raw_data sheet
    def raw_data_gen(df, calc_date):
        calc_df = df[df['Date'] == calc_date].copy()
        calc_df.sort_values(by='AveragePrice', ascending=False, inplace=True)
        return calc_df[['Date', 'region', 'AveragePrice', 'Total Volume']]

    raw_data_rows = dataframe_to_rows(
        raw_data_gen(
            df,
            calc_date),
        index=False,
        header=True)  # False index

    for r_idx, row in enumerate(raw_data_rows, 1):
        for c_idx, value in enumerate(row, 1):
            raw_data_sheet.cell(row=r_idx, column=c_idx, value=value)


def generate_final_report_job(df, report_template, calc_date_list):
    global date_index
    calc_date = calc_date_list[date_index]
    wb = load_workbook(report_template)  # workbook
    report_sheet = wb['report']
    raw_data_sheet = wb['raw']
    raw_sheet_update(raw_data_sheet, df, calc_date)
    report_data = report_data_gen(df, calc_date)
    report_sheet_update(report_sheet, report_data, raw_data_sheet)
    calc_date = pd.to_datetime(str(calc_date))
    timestr = calc_date.strftime('%Y_%m_%d_%H_%M_%S')
    print(timestr)
    file_path = f'report_{timestr}.xlsx'
    wb.save(file_path)
    date_index += 1
from datetime import datetime

if __name__ == "__main__":
    file = 'avocado.csv'
    report_template = 'report_template.xlsx'
    df = pd.read_csv(file, parse_dates=['Date'])
    date_index = 0
    calc_date_list = df['Date'].sort_values().unique()
    start_time = datetime.now()f
    schedule.every(5).seconds.do(
        generate_final_report_job,
        df=df,
        report_template=report_template,
        calc_date_list=calc_date_list)
    while True:
        schedule.run_pending()
        time.sleep(1)


